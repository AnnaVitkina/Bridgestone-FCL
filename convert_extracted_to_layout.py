from __future__ import annotations

from pathlib import Path
from typing import Optional
import re
from copy import copy

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

PROCESSING_DIR = Path(__file__).resolve().parent / "processing"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"


def ask_user_for_source_file(files: list[Path]) -> Path:
    """Ask user which extracted file should be converted."""
    print("\nExtracted files in processing:")
    for idx, file_path in enumerate(files, start=1):
        print(f"  {idx}. {file_path.name}")

    while True:
        raw_value = input("Enter file number to convert: ").strip()
        if raw_value.isdigit():
            selected_idx = int(raw_value)
            if 1 <= selected_idx <= len(files):
                return files[selected_idx - 1]
        print("Invalid choice. Enter a valid number.")


def split_route_and_rate_columns(df: pd.DataFrame) -> tuple[list[str], list[str], Optional[str]]:
    """Split dataframe columns into route details and rate details."""
    currency_col = next(
        (col for col in df.columns if str(col).strip().lower() == "currency"), None
    )

    rate_columns: list[str] = []
    for col in df.columns:
        col_name = str(col).strip().lower()
        if currency_col and col == currency_col:
            continue
        if re.search(r"\d", col_name) or "rate" in col_name or "cost" in col_name:
            rate_columns.append(col)

    if not rate_columns:
        numeric_like_cols: list[str] = []
        for col in df.columns:
            if currency_col and col == currency_col:
                continue
            converted = pd.to_numeric(df[col], errors="coerce")
            if converted.notna().mean() > 0.8:
                numeric_like_cols.append(col)
        rate_columns = numeric_like_cols

    route_columns = [col for col in df.columns if col not in rate_columns and col != currency_col]
    return route_columns, rate_columns, currency_col


def add_multi_picking_columns(df: pd.DataFrame, route_columns: list[str]) -> tuple[pd.DataFrame, list[str]]:
    """Add Second load and Multi-picking FOB columns after Origin City."""
    origin_city_col = next(
        (col for col in route_columns if str(col).strip().lower() == "origin city"), None
    )
    if origin_city_col is None:
        return df, route_columns

    prepared_df = df.copy()
    has_plus = prepared_df[origin_city_col].apply(
        lambda value: isinstance(value, str) and "+" in value
    )
    if not has_plus.any():
        return prepared_df, route_columns

    prepared_df["Second load"] = has_plus.map({True: "YES", False: "No"})
    prepared_df["Multi-picking FOB"] = prepared_df[origin_city_col].where(has_plus, "")

    origin_idx = route_columns.index(origin_city_col)
    new_route_columns = (
        route_columns[: origin_idx + 1]
        + ["Second load", "Multi-picking FOB"]
        + route_columns[origin_idx + 1 :]
    )
    return prepared_df, new_route_columns


def clean_postal_code_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only digits in postal code columns, if they exist."""
    cleaned_df = df.copy()

    def normalize_postal_code(value: object) -> object:
        if pd.isna(value):
            return value
        # Preserve leading zeros when value is already text (e.g. "09007").
        if isinstance(value, str):
            return re.sub(r"\D", "", value)
        # Handle numeric cells safely without introducing artifacts.
        if isinstance(value, (int, float)):
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return re.sub(r"\D", "", str(value))
        return re.sub(r"\D", "", str(value))

    for col in cleaned_df.columns:
        col_name = str(col).strip().lower()
        if col_name in {"origin postal code", "destination postal code"}:
            cleaned_df[col] = cleaned_df[col].apply(normalize_postal_code)
    return cleaned_df


def detect_primary_carrier(df: pd.DataFrame) -> Optional[str]:
    """Return the main carrier name from the extracted file."""
    carrier_col = next(
        (col for col in df.columns if str(col).strip().lower() == "carrier name in control pay"),
        None,
    )
    if carrier_col is None:
        return None
    carriers = df[carrier_col].dropna().astype(str).map(str.strip)
    carriers = carriers[carriers != ""]
    if carriers.empty:
        return None
    return carriers.value_counts().index[0]


def get_rate_layout_config(rate_columns: list[str]) -> list[tuple[str, str, str]]:
    """
    Build per-rate layout tuple: (source_rate_col, display_rate_col, applies_if_text).
    Base behavior keeps original rate name and generic applies-if label.
    """
    global_rules = {
        "20'": ("20'", "Equipment Type equals '13/TC 20\" Dry 30 M3'"),
        "40'": ("40'", "Equipment Type equals '09/TC 40\" Dry 60 M3'"),
        "40' HQ": ("40'HC", "Equipment Type equals '10/TC 40\" H-Cube 76 M3'"),
    }

    configured_layout: list[tuple[str, str, str]] = []
    for rate_col in rate_columns:
        if rate_col in global_rules:
            display_name, applies_if = global_rules[rate_col]
            configured_layout.append((rate_col, display_name, applies_if))
        else:
            configured_layout.append((rate_col, rate_col, "Apply if"))
    return configured_layout


def _normalize_multi_picking_value(value: str) -> str:
    """Convert value to uppercase underscore format."""
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", value.strip().upper())
    return normalized.strip("_")


def _city_value_variants(value: str) -> list[str]:
    """Return city value with optional title-case variant for uppercase names."""
    base = value.strip()
    variants = [base]
    has_letters = bool(re.search(r"[A-Za-z]", base))
    has_digits = bool(re.search(r"\d", base))
    if has_letters and not has_digits and base == base.upper():
        title_variant = base.title()
        if title_variant != base:
            variants.append(title_variant)
    return variants


def _normalize_city_key(value: str) -> str:
    """Normalize city string for robust override matching."""
    return re.sub(r"[^A-Z0-9]+", " ", value.upper()).strip()


def _resolve_destination_override(
    carrier_name: Optional[str], city_name: str, common_values: list[str]
) -> Optional[list[str]]:
    """Return carrier-specific destination override values when configured."""
    normalized_city = _normalize_city_key(city_name)
    normalized_carrier = (carrier_name or "").strip().lower()

    if normalized_carrier == "ceva freight poland":
        ceva_destination_overrides = {
            "SKIKDA": ["Skikda", "OUM EL BOUAGHI"],
            "BUENOS AIRES": ["Llavallol", "Buenos Aires"],
        }
        return ceva_destination_overrides.get(normalized_city)

    if normalized_carrier == "gerco":
        gerco_override_replace = {
            "SINT PETERSBURG": ["Sint Petersburg", "Ulyanovsk"],
            "BEIRUT": ["Beirut", "MKALLES", "Mkalles", "Lebanon"],
            "DAMMAM": ["Dammam", "AL KHOBAR"],
            "AQABA": ["Aqaba", "AMMAN"],
            "TUNIS": ["Tunis", "RADES"],
            "SKIKDA": ["Skikda", "OUM EL BOUAGHI"],
            "BUENOS AIRES": ["Llavallol", "Buenos Aires"],
            "MOMBASA": ["Mombasa", "Monbassa"],
            "MONBASSA": ["Mombasa", "Monbassa"],
        }
        gerco_override_with_common = {
            "ASHDOD PORT": ["ASHDOD"],
            "PORT ELIZABETH": ["PORT ELIZABETH"],
        }

        if "JEBEL ALI" in normalized_city:
            return ["Jebel Ali", "Dubai", "P.O BOX 21541", "B1", "13TH"]

        for key, values in gerco_override_replace.items():
            if normalized_city.startswith(key):
                return values

        for key, values in gerco_override_with_common.items():
            if normalized_city.startswith(key):
                merged = values + [value for value in common_values if value not in values]
                return merged

    if normalized_carrier == "dolt logistics":
        dolt_overrides = {
            "TINCAN ISLAND": ["TINCAN", "TIN CAN"],
            "BUENOS AIRES": ["Llavallol", "Buenos Aires"],
            "ASHDOD PORT": ["ASHDOD"],
            "ASHDOD TEL AVIV": ["Ashdod", "TEL-AVIV", "TEL AVIV", "TELAVIV"],
        }
        for key, values in dolt_overrides.items():
            if normalized_city.startswith(key):
                return values

    return None


def _get_incoterm_rules(
    carrier_name: Optional[str], df: pd.DataFrame
) -> list[tuple[str, str, str, str]]:
    """Return carrier-specific incoterm condition rules."""
    normalized_carrier = (carrier_name or "").strip().lower()
    incoterm_col = next(
        (col for col in df.columns if str(col).strip().lower() == "incoterm"),
        None,
    )
    if incoterm_col is None:
        return []

    present_incoterms = {
        value.strip().upper()
        for value in df[incoterm_col].dropna().astype(str)
        if value.strip()
    }

    def rule_if_present(
        key: str, operator: str, values: str, scope: str
    ) -> Optional[tuple[str, str, str, str]]:
        return (key, operator, values, scope) if key.upper() in present_incoterms else None

    if normalized_carrier == "ceva freight poland":
        candidate_rules = [
            rule_if_present("DAT", "equals", "DAT, DPU", "all items"),
            rule_if_present("CFR", "equals", "CFR, CIF", "all items"),
        ]
        return [rule for rule in candidate_rules if rule is not None]
    if normalized_carrier == "gerco":
        candidate_rules = [
            rule_if_present("DAT", "equals", "DAT, DPU", "all items"),
            rule_if_present("CFR", "equals", "CFR, CIF", "all items"),
        ]
        return [rule for rule in candidate_rules if rule is not None]
    if normalized_carrier == "dhl global forwarding hungary ltd":
        candidate_rules = [
            rule_if_present("DAT", "equals", "DAT, DPU", "all items"),
            rule_if_present("CFR", "equals", "CFR, CIF", "all items"),
        ]
        return [rule for rule in candidate_rules if rule is not None]
    if normalized_carrier == "dolt logistics":
        candidate_rules = [
            rule_if_present("DPU", "equals", "DAT, DPU", "all items"),
            rule_if_present("DAT", "equals", "DAT, DPU", "all items"),
            rule_if_present("CFR", "equals", "CFR, CIF", "all items"),
        ]
        return [rule for rule in candidate_rules if rule is not None]
    return []


def build_conditions_rows(df: pd.DataFrame, carrier_name: Optional[str]) -> list[list[object]]:
    """Build rules for Destination City and Multi-picking FOB."""
    rows: list[list[object]] = []
    rule_no = 1

    destination_col = next(
        (col for col in df.columns if str(col).strip().lower() == "destination city"),
        None,
    )
    if destination_col is not None:
        rows.append(["", "Destination City", "", "", ""])
        destination_values_all = (
            df[destination_col]
            .dropna()
            .astype(str)
            .map(str.strip)
            .drop_duplicates()
        )
        destination_values = destination_values_all

        for city_value in destination_values:
            # Split only by backslash separators used between city segments.
            # Keep tokens like "W/H" intact (do not split by forward slash).
            city_parts = [
                part.strip() for part in re.split(r"\s*\\\s*", city_value) if part.strip()
            ]

            # Add compact tokens used in conditions (e.g. "13TH", "B1").
            for part in list(city_parts):
                for token in re.findall(r"\b(?:\d+[A-Z]{2}|[A-Z]\d+)\b", part.upper()):
                    if token not in city_parts:
                        city_parts.append(token)

            values_with_case_variants: list[str] = []
            for part in city_parts:
                for variant in _city_value_variants(part):
                    if variant not in values_with_case_variants:
                        values_with_case_variants.append(variant)

            override_values = _resolve_destination_override(
                carrier_name, city_value, values_with_case_variants
            )
            if override_values:
                values_with_case_variants = override_values
            rows.append(
                [
                    rule_no,
                    city_value,
                    "contains",
                    "; ".join(values_with_case_variants),
                    "all items",
                ]
            )
            rule_no += 1

    multi_picking_col = next(
        (col for col in df.columns if str(col).strip().lower() == "multi-picking fob"),
        None,
    )
    if multi_picking_col is not None:
        if rows:
            rows.append(["", "", "", "", ""])
        rows.append(["", "Multi-picking FOB", "", "", ""])
        multi_values = (
            df[multi_picking_col]
            .dropna()
            .astype(str)
            .map(str.strip)
            .loc[lambda s: s != ""]
            .drop_duplicates()
        )
        for multi_value in multi_values:
            rows.append(
                [
                    rule_no,
                    multi_value,
                    "equals",
                    _normalize_multi_picking_value(multi_value),
                    "in all items",
                ]
            )
            rule_no += 1

    second_load_col = next(
        (col for col in df.columns if str(col).strip().lower() == "second load"),
        None,
    )
    if second_load_col is not None:
        if rows:
            rows.append(["", "", "", "", ""])
        rows.append(["", "Second load", "", "", ""])
        rows.append([rule_no, "NO", "does not equal to", "YES", "in all items"])
        rule_no += 1

    normalized_carrier = (carrier_name or "").strip().lower()
    if normalized_carrier == "dhl global forwarding hungary ltd":
        if rows:
            rows.append(["", "", "", "", ""])
        rows.append(["", "Origin Postal Code", "", "", ""])
        rows.append([rule_no, "Tatabanya", "starts with", "2800, 2851", "in all items"])
        rule_no += 1
    if normalized_carrier == "dolt logistics":
        if rows:
            rows.append(["", "", "", "", ""])
        rows.append(["", "Origin Postal Code", "", "", ""])
        rows.append([rule_no, "Puente S. Miguel", "starts with", "39380, 39530", "in all items"])
        rule_no += 1
        rows.append([rule_no, "9006/09006", "starts with", "09006", "in all items"])
        rule_no += 1

    incoterm_rules = _get_incoterm_rules(carrier_name, df)
    if incoterm_rules:
        if rows:
            rows.append(["", "", "", "", ""])
        rows.append(["", "Incoterm", "", "", ""])
        for name, operator, values, scope in incoterm_rules:
            rows.append([rule_no, name, operator, values, scope])
            rule_no += 1

    return rows


def _split_rule_values(values: object) -> list[str]:
    """Split condition values by supported delimiters."""
    if not isinstance(values, str):
        return []
    return [part.strip() for part in re.split(r"[;,]", values) if part.strip()]


def _normalize_for_compare(value: object) -> str:
    """Normalize text for case-insensitive comparisons."""
    if value is None:
        return ""
    return str(value).strip().upper()


def _build_condition_rules_by_section(
    conditions_rows: list[list[object]],
) -> dict[str, list[dict[str, object]]]:
    """Build section-to-rules map from the Conditions rows payload."""
    rules_by_section: dict[str, list[dict[str, object]]] = {}
    current_section: Optional[str] = None

    for row in conditions_rows:
        rule_no, name, operator, values, _scope = row
        name_text = str(name).strip() if isinstance(name, str) else ""
        operator_text = str(operator).strip().lower() if isinstance(operator, str) else ""

        is_section_header = (
            (rule_no == "" or rule_no is None)
            and name_text != ""
            and operator_text == ""
        )
        if is_section_header:
            current_section = name_text.lower()
            rules_by_section.setdefault(current_section, [])
            continue

        is_rule_row = current_section is not None and name_text != "" and operator_text != ""
        if is_rule_row:
            rules_by_section[current_section].append(
                {
                    "name": name_text,
                    "operator": operator_text,
                    "values": _split_rule_values(values),
                }
            )

    return rules_by_section


def apply_condition_rules_to_df(
    df: pd.DataFrame, conditions_rows: list[list[object]]
) -> pd.DataFrame:
    """
    Apply condition logic to dataframe values:
    if a value matches a rule's Values, replace it with that rule Name.
    """
    updated_df = df.copy()
    rules_by_section = _build_condition_rules_by_section(conditions_rows)

    for section_name, rules in rules_by_section.items():
        target_column = next(
            (col for col in updated_df.columns if str(col).strip().lower() == section_name),
            None,
        )
        if target_column is None or not rules:
            continue

        def map_value(original_value: object) -> object:
            raw_text = str(original_value).strip() if original_value is not None else ""
            if raw_text == "":
                return original_value
            normalized_value = _normalize_for_compare(raw_text)

            for rule in rules:
                rule_name = str(rule["name"])
                rule_operator = str(rule["operator"])
                rule_values = [_normalize_for_compare(v) for v in rule["values"]]

                # Keep existing canonical value as-is.
                if normalized_value == _normalize_for_compare(rule_name):
                    return rule_name

                if rule_operator == "equals":
                    if normalized_value in rule_values:
                        return rule_name
                elif rule_operator == "starts with":
                    normalized_value_no_zeros = normalized_value.lstrip("0")
                    if any(
                        normalized_value.startswith(prefix)
                        or normalized_value_no_zeros.startswith(prefix.lstrip("0"))
                        for prefix in rule_values
                    ):
                        return rule_name
                elif rule_operator == "contains":
                    if any(token and token in normalized_value for token in rule_values):
                        return rule_name

            return original_value

        updated_df[target_column] = updated_df[target_column].apply(map_value)

    return updated_df


def write_conditions_sheet(
    wb: Workbook, conditions_rows: list[list[object]]
) -> None:
    """Create Conditions tab with generated rule rows."""
    ws = wb.create_sheet(title="Conditions")
    headers = ["Rule #", "Name", "Operator", "Values", "Scope"]
    ws.append(headers)

    for row in conditions_rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True, size=9)
    section_font = Font(bold=True, size=9)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(headers) + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.border = border
        ws.column_dimensions[header_cell.column_letter].width = 28 if col_idx in (2, 4) else 16

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        if ws.cell(row=row_idx, column=1).value == "" and ws.cell(row=row_idx, column=2).value in {
            "Destination City",
            "Multi-picking FOB",
            "Second load",
            "Origin Postal Code",
            "Incoterm",
        }:
            ws.cell(row=row_idx, column=2).font = section_font

    ws.freeze_panes = "A2"


def highlight_condition_values(
    converted_ws,
    conditions_rows: list[list[object]],
    data_start_row: int,
    header_row: int,
    max_col: int,
) -> None:
    """Underline and grey-fill Converted cells for matching condition names per column."""
    rules_by_section = _build_condition_rules_by_section(conditions_rows)
    highlight_names_by_section = {
        section: {str(rule["name"]).strip().casefold() for rule in rules}
        for section, rules in rules_by_section.items()
    }

    highlight_fill = PatternFill("solid", fgColor="D9D9D9")

    for col_idx in range(1, max_col + 1):
        header_value = converted_ws.cell(row=header_row, column=col_idx).value
        header_key = str(header_value).strip().lower() if isinstance(header_value, str) else ""
        highlight_names = highlight_names_by_section.get(header_key, set())
        if not highlight_names:
            continue

        for row_idx in range(data_start_row, converted_ws.max_row + 1):
            cell = converted_ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, str) and cell.value.strip().casefold() in highlight_names:
                cell.fill = highlight_fill
                underlined_font = copy(cell.font)
                underlined_font.underline = "single"
                cell.font = underlined_font


def write_layout_xlsx(df: pd.DataFrame, output_path: Path) -> None:
    """Write dataframe to target layout with Route details and Rates details sections."""
    df = clean_postal_code_columns(df)
    route_columns, rate_columns, currency_col = split_route_and_rate_columns(df)
    if not rate_columns:
        raise ValueError("Could not identify rate columns from extracted file.")
    df, route_columns = add_multi_picking_columns(df, route_columns)
    carrier_name = detect_primary_carrier(df)
    rate_layout = get_rate_layout_config(rate_columns)
    conditions_rows = build_conditions_rows(df, carrier_name)
    df = apply_condition_rules_to_df(df, conditions_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Converted"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    title_font = Font(bold=True, size=10)
    header_font = Font(bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    route_start = 1
    route_end = len(route_columns)
    rate_start = route_end + 1
    rate_end = rate_start + (2 * len(rate_layout)) - 1

    ws.merge_cells(start_row=1, start_column=route_start, end_row=1, end_column=route_end)

    for idx, (_, display_rate_col, applies_if_text) in enumerate(rate_layout):
        col_start = rate_start + idx * 2
        col_end = col_start + 1
        ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        ws.cell(row=1, column=col_start, value=f"Transport cost ({display_rate_col})")
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        ws.cell(row=2, column=col_start, value=applies_if_text)
        ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
        ws.cell(row=3, column=col_start, value="Rate by: per shipment")

    for col_idx, route_col in enumerate(route_columns, start=1):
        ws.cell(row=4, column=col_idx, value=route_col)

    for idx in range(len(rate_layout)):
        col_start = rate_start + idx * 2
        ws.cell(row=4, column=col_start, value="Currency")
        ws.cell(row=4, column=col_start + 1, value="Flat")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=5):
        for col_idx, route_col in enumerate(route_columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(route_col))

        for idx, (source_rate_col, _, _) in enumerate(rate_layout):
            col_start = rate_start + idx * 2
            rate_value = row.get(source_rate_col)
            currency_value = row.get(currency_col) if currency_col else None
            if currency_value is None or (isinstance(currency_value, float) and pd.isna(currency_value)):
                currency_value = ""
            ws.cell(row=row_idx, column=col_start, value=currency_value)
            ws.cell(row=row_idx, column=col_start + 1, value=rate_value)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=rate_end):
        for cell in row:
            cell.border = border
            if cell.row <= 4:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

    for idx in range(len(rate_layout)):
        ws.cell(row=1, column=rate_start + idx * 2).font = title_font

    for col_idx in range(1, rate_end + 1):
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = 18

    ws.freeze_panes = "A5"
    write_conditions_sheet(wb, conditions_rows)
    highlight_condition_values(
        ws, conditions_rows, data_start_row=5, header_row=4, max_col=rate_end
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    files = sorted(PROCESSING_DIR.glob("*_extracted.xlsx"))
    if not files:
        raise FileNotFoundError(f"No extracted files found in {PROCESSING_DIR}")

    source_file = ask_user_for_source_file(files)
    df = pd.read_excel(
        source_file,
        dtype={
            "Origin Postal Code": "string",
            "Destination Postal Code": "string",
        },
    )

    output_file = OUTPUT_DIR / f"{source_file.stem}_layout.xlsx"
    write_layout_xlsx(df, output_file)
    print(f"Converted file saved to: {output_file}")


if __name__ == "__main__":
    main()
