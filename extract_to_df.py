from __future__ import annotations

from pathlib import Path
from typing import Optional
from datetime import datetime
import re

import pandas as pd
from openpyxl import load_workbook

INPUT_DIR = Path(__file__).resolve().parent / "input"
OUTPUT_DIR = Path(__file__).resolve().parent / "processing"
DEFAULT_SHEET = "Sea FCL - Rate card"


def ask_user_for_sheet(file_path: Path, available_sheets: list[str]) -> Optional[str]:
    """Prompt the user to select a sheet in terminal when default is missing."""
    print(f"\nDefault sheet '{DEFAULT_SHEET}' was not found in: {file_path.name}")
    print("Available sheets:")
    for idx, sheet in enumerate(available_sheets, start=1):
        print(f"  {idx}. {sheet}")

    while True:
        raw_value = input(
            "Enter sheet number to convert (or press Enter to skip this file): "
        ).strip()
        if raw_value == "":
            return None
        if raw_value.isdigit():
            selected_idx = int(raw_value)
            if 1 <= selected_idx <= len(available_sheets):
                return available_sheets[selected_idx - 1]
        print("Invalid choice. Please enter a valid sheet number.")


def ask_user_for_files(files: list[Path]) -> list[Path]:
    """Prompt the user to choose one file or all files."""
    print("\nAvailable files:")
    for idx, file_path in enumerate(files, start=1):
        print(f"  {idx}. {file_path.name}")

    while True:
        raw_value = input(
            "Enter file number to convert, or 'a' for all files: "
        ).strip().lower()
        if raw_value == "a":
            return files
        if raw_value.isdigit():
            selected_idx = int(raw_value)
            if 1 <= selected_idx <= len(files):
                return [files[selected_idx - 1]]
        print("Invalid choice. Enter a valid number or 'a'.")


def clean_rate_card_df(df: pd.DataFrame) -> pd.DataFrame:
    """Apply lightweight cleaning to extracted data."""
    cleaned_df = df.copy()
    cleaned_df.columns = [str(col).strip() for col in cleaned_df.columns]

    for col in cleaned_df.select_dtypes(include=["object", "string"]).columns:
        cleaned_df[col] = cleaned_df[col].apply(
            lambda value: value.strip() if isinstance(value, str) else value
        )

    cleaned_df = cleaned_df.dropna(axis=0, how="all")
    cleaned_df = cleaned_df.dropna(axis=1, how="all")

    if "Unnamed: 0" in cleaned_df.columns:
        cleaned_df = cleaned_df.rename(columns={"Unnamed: 0": "row_label"})

    return cleaned_df.reset_index(drop=True)


def _extract_currency_from_row(row: pd.Series) -> Optional[str]:
    """Extract a currency code from a sheet row."""
    candidates = []
    for value in row.tolist():
        if isinstance(value, str):
            token = value.strip().upper()
            if len(token) == 3 and token.isalpha():
                candidates.append(token)
    if not candidates:
        return None
    return max(set(candidates), key=candidates.count)


def _extract_decimal_places(number_format: str) -> Optional[int]:
    """Infer visible decimal places from an Excel number format."""
    if not isinstance(number_format, str) or "." not in number_format:
        return None
    first_section = number_format.split(";")[0]
    match = re.search(r"\.([0#]+)", first_section)
    if not match:
        return None
    return len(match.group(1))


def _format_visible_cell_value(value: object, number_format: str) -> object:
    """
    Convert numeric values to their visible precision in Excel.
    Example: 2377.30387982 with 2-decimal format -> "2377.30".
    """
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        decimal_places = _extract_decimal_places(number_format)
        if decimal_places is None:
            return value
        rounded = round(float(value), decimal_places)
        return f"{rounded:.{decimal_places}f}"
    return value


def parse_rate_card_sheet(file_path: Path, sheet_name: str) -> tuple[pd.DataFrame, Optional[str]]:
    """Parse the rate card sheet by locating its real header row."""
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name]

    raw_rows: list[list[object]] = []
    for row in worksheet.iter_rows():
        raw_rows.append(
            [_format_visible_cell_value(cell.value, cell.number_format) for cell in row]
        )
    workbook.close()

    raw_df = pd.DataFrame(raw_rows)
    header_row_idx: Optional[int] = None

    for idx, row in raw_df.iterrows():
        normalized = [str(value).strip().lower() for value in row.tolist()]
        if "rc#" in normalized and "lane id" in normalized:
            header_row_idx = idx
            break

    if header_row_idx is None:
        raise ValueError(f"Could not find header row in {file_path.name} / {sheet_name}")

    currency = None
    if header_row_idx > 0:
        currency = _extract_currency_from_row(raw_df.iloc[header_row_idx - 1])

    headers = raw_df.iloc[header_row_idx].tolist()
    normalized_headers: list[str] = []
    for idx, header in enumerate(headers, start=1):
        if pd.isna(header):
            normalized_headers.append(f"column_{idx}")
        else:
            text = str(header).strip()
            normalized_headers.append(text if text else f"column_{idx}")

    data_df = raw_df.iloc[header_row_idx + 1 :].copy()
    data_df.columns = normalized_headers
    return data_df, currency


def load_rate_card_df(
    input_dir: Path = INPUT_DIR,
) -> tuple[pd.DataFrame, list[Path], dict[str, str]]:
    """
    Load and combine all xlsx files from input_dir into a single DataFrame.
    Uses DEFAULT_SHEET when available; otherwise asks user in terminal.
    """
    files = sorted(input_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {input_dir}")

    selected_files = ask_user_for_files(files)
    dataframes: list[pd.DataFrame] = []
    currencies_by_file: dict[str, str] = {}

    for file_path in selected_files:
        xls = pd.ExcelFile(file_path)

        if DEFAULT_SHEET in xls.sheet_names:
            sheet_to_read = DEFAULT_SHEET
        else:
            sheet_to_read = ask_user_for_sheet(file_path, xls.sheet_names)
            if sheet_to_read is None:
                print(f"Skipping file: {file_path.name}")
                continue

        df, currency = parse_rate_card_sheet(file_path, sheet_to_read)
        df = clean_rate_card_df(df)
        if currency:
            currencies_by_file[file_path.name] = currency
            df["currency"] = currency
        dataframes.append(df)

    if not dataframes:
        raise ValueError("No files were loaded into DataFrame.")

    return pd.concat(dataframes, ignore_index=True), selected_files, currencies_by_file


def save_df_to_excel(
    df: pd.DataFrame, selected_files: list[Path], output_dir: Path = OUTPUT_DIR
) -> Path:
    """Save dataframe to an Excel file in processing folder."""
    output_dir.mkdir(parents=True, exist_ok=True)
    if len(selected_files) == 1:
        output_name = f"{selected_files[0].stem}_extracted.xlsx"
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"fcl_rate_card_combined_{timestamp}.xlsx"
    output_path = output_dir / output_name
    df.to_excel(output_path, index=False)
    return output_path


if __name__ == "__main__":
    df, selected_files, currencies_by_file = load_rate_card_df()
    saved_path = save_df_to_excel(df, selected_files)
    print(f"Loaded DataFrame shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    if len(selected_files) == 1:
        currency = currencies_by_file.get(selected_files[0].name, "UNKNOWN")
        print(f"currency = {currency}")
    else:
        print(f"Currencies by file: {currencies_by_file}")
    print(f"Saved to: {saved_path}")
